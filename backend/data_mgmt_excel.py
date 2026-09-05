"""Template Excel, ekspor master, dan pembaca berkas impor (openpyxl).

Tata letak setiap sheet data:
  baris 1 = kunci kolom teknis (dibaca mesin), baris 2 = label + aturan (untuk manusia),
  baris 3.. = data. Sheet 'PETUNJUK' dan 'DAFTAR NILAI' tidak diimpor.
"""
import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from data_mgmt_schema import ENTITIES, ENTITY_BY_SHEET, enum_groups_used, enum_options

HEADER_ROW, DESC_ROW, FIRST_DATA_ROW = 1, 2, 3
GUIDE_SHEET, VALUES_SHEET = "PETUNJUK", "DAFTAR NILAI"

_HEAD_FILL = PatternFill("solid", fgColor="0F766E")
_REQ_FILL = PatternFill("solid", fgColor="115E59")
_DESC_FILL = PatternFill("solid", fgColor="F0FDFA")
_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TYPE_LABEL = {"str": "teks", "int": "angka bulat", "float": "angka", "bool": "TRUE/FALSE",
              "enum": "pilihan", "list": "daftar (pisah ;)", "email": "email", "phone": "no. HP"}


def _rule_text(f: dict) -> str:
    parts = ["WAJIB" if f["required"] else "opsional", TYPE_LABEL.get(f["type"], f["type"])]
    if f["enum"]:
        parts.append("lihat DAFTAR NILAI: " + f["enum"])
    if f.get("default") not in (None, ""):
        parts.append(f"bawaan {f['default']}")
    if f["desc"]:
        parts.append(f["desc"])
    if f["example"] not in (None, ""):
        parts.append(f"contoh: {f['example']}")
    return " · ".join(parts)


def _write_guide(wb: Workbook):
    ws = wb.active
    ws.title = GUIDE_SHEET
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 90
    ws["B1"] = "TEMPLATE MIGRASI MASTER DATA — SIPRO"
    ws["B1"].font = Font(bold=True, size=14, color="0F766E")
    ws["B2"] = f"Dibuat {datetime.now():%d %b %Y %H:%M}. Isi sheet sesuai urutan di bawah, " \
               "lalu unggah di menu Admin → Manajemen Data → Migrasi Excel."
    rules = [
        "Baris 1 setiap sheet = kunci kolom (JANGAN diubah/dihapus). Baris 2 = keterangan & aturan.",
        "Isi data mulai baris 3. Baris kosong diabaikan. Kolom opsional boleh dikosongkan.",
        "Kolom bertipe 'pilihan' hanya menerima nilai kanonik di sheet DAFTAR NILAI (kolom Nilai).",
        "Kolom TRUE/FALSE juga menerima Ya/Tidak, 1/0.",
        "Kolom rujukan (mis. Kode proyek) harus cocok dengan kode di sheet rujukannya atau data "
        "yang sudah ada di sistem.",
        "Sheet yang tidak diperlukan boleh dikosongkan (hanya baris 1-2) atau dihapus.",
        "Saat impor, sistem memvalidasi dulu (pratinjau) — tidak ada yang ditulis sampai Anda "
        "menekan 'Jalankan impor'. Pilih perilaku duplikat: perbarui (upsert) atau lewati.",
    ]
    r = 4
    ws.cell(row=r, column=2, value="Aturan pengisian").font = Font(bold=True)
    for i, t in enumerate(rules, start=1):
        ws.cell(row=r + i, column=2, value=f"{i}.")
        c = ws.cell(row=r + i, column=3, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r = r + len(rules) + 2
    ws.cell(row=r, column=2, value="Urutan sheet & isinya").font = Font(bold=True)
    for i, e in enumerate(ENTITIES, start=1):
        ws.cell(row=r + i, column=2, value=f"{i}. {e['sheet']}").font = Font(bold=True)
        ws.cell(row=r + i, column=3, value=f"{e['desc']} Kunci duplikat: "
                + " + ".join(e["key_fields"])).alignment = Alignment(wrap_text=True)


def _write_values(wb: Workbook) -> dict:
    """Sheet DAFTAR NILAI; kembalikan {group: 'range'} untuk validasi dropdown."""
    ws = wb.create_sheet(VALUES_SHEET)
    ranges, col = {}, 1
    for group in enum_groups_used():
        opts = enum_options(group)
        head = ws.cell(row=1, column=col, value=f"Nilai ({group})")
        head.font = Font(bold=True, color="FFFFFF")
        head.fill = _HEAD_FILL
        ws.cell(row=1, column=col + 1, value="Arti").font = Font(bold=True)
        for i, (v, lbl) in enumerate(opts.items(), start=2):
            ws.cell(row=i, column=col, value=v)
            ws.cell(row=i, column=col + 1, value=lbl)
        letter = get_column_letter(col)
        ranges[group] = f"'{VALUES_SHEET}'!${letter}$2:${letter}${len(opts) + 1}"
        ws.column_dimensions[letter].width = max(16, max(len(v) for v in opts) + 2)
        ws.column_dimensions[get_column_letter(col + 1)].width = 28
        col += 3
    return ranges


def _write_entity_sheet(wb: Workbook, ent: dict, rows: list, ranges: dict):
    ws = wb.create_sheet(ent["sheet"])
    for ci, f in enumerate(ent["fields"], start=1):
        h = ws.cell(row=HEADER_ROW, column=ci, value=f["key"])
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = _REQ_FILL if f["required"] else _HEAD_FILL
        h.border = _BORDER
        d = ws.cell(row=DESC_ROW, column=ci, value=f"{f['label']}\n{_rule_text(f)}")
        d.font = Font(italic=True, size=9, color="475569")
        d.fill = _DESC_FILL
        d.alignment = Alignment(wrap_text=True, vertical="top")
        d.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = f.get("width") or max(
            16, min(40, len(f["label"]) + 6))
        if f["enum"] and f["enum"] in ranges:
            dv = DataValidation(type="list", formula1=ranges[f["enum"]], allow_blank=True,
                                showErrorMessage=True, errorTitle="Nilai tidak dikenal",
                                error="Pilih nilai dari sheet DAFTAR NILAI.")
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(ci)}{FIRST_DATA_ROW}:{get_column_letter(ci)}2000")
        elif f["type"] == "bool":
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(ci)}{FIRST_DATA_ROW}:{get_column_letter(ci)}2000")
    ws.row_dimensions[DESC_ROW].height = 64
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=1)
    for ri, row in enumerate(rows, start=FIRST_DATA_ROW):
        for ci, f in enumerate(ent["fields"], start=1):
            v = row.get(f["key"])
            if isinstance(v, list):
                v = "; ".join(str(x) for x in v)
            elif isinstance(v, bool):
                v = "TRUE" if v else "FALSE"
            elif isinstance(v, dict):
                v = None
            ws.cell(row=ri, column=ci, value=v)


def build_workbook(data: dict, with_example: bool = False) -> bytes:
    """data = {entity_key: [row dict]}; kosong → template. with_example → 1 baris contoh."""
    wb = Workbook()
    _write_guide(wb)
    ranges = _write_values(wb)
    for ent in ENTITIES:
        rows = data.get(ent["key"]) or []
        if not rows and with_example:
            rows = [{f["key"]: f["example"] for f in ent["fields"] if f["example"] not in ("", None)}]
        _write_entity_sheet(wb, ent, rows, ranges)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def parse_workbook(content: bytes) -> dict:
    """→ {"sheets": {entity_key: [{"row": n, **values}]}, "unknown_sheets": [...]}"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out, unknown = {}, []
    for ws in wb.worksheets:
        title = ws.title.strip()
        if title.upper() in (GUIDE_SHEET, VALUES_SHEET):
            continue
        ent = ENTITY_BY_SHEET.get(title.lower())
        if not ent:
            unknown.append(title)
            continue
        label_to_key = {f["label"].lower(): f["key"] for f in ent["fields"]}
        valid_keys = {f["key"] for f in ent["fields"]}
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            out[ent["key"]] = []
            continue
        cols = []
        for h in header:
            h = (str(h).strip() if h is not None else "")
            key = h if h in valid_keys else label_to_key.get(h.lower())
            cols.append(key)
        rows = []
        for ri, raw in enumerate(rows_iter, start=2):
            if ri == DESC_ROW:
                continue
            vals = {}
            for key, v in zip(cols, raw):
                if key:
                    vals[key] = _cell_value(v)
            if any(v is not None for v in vals.values()):
                vals["row"] = ri
                rows.append(vals)
        out[ent["key"]] = rows
    wb.close()
    return {"sheets": out, "unknown_sheets": unknown}
