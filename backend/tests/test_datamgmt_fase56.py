"""Backend tests for Manajemen Data (Fase Data Management)."""
import io
import json
import os
import time

import openpyxl
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/") or (
    open("/app/frontend/.env").read().split("REACT_APP_BACKEND_URL=")[1].split()[0]
)
API = f"{BASE_URL}/api"
PASSWORD = "Sipro#2026"


def _login(email):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": PASSWORD}, timeout=30)
    assert r.status_code == 200, f"login failed for {email}: {r.status_code} {r.text}"
    data = r.json()
    token = data.get("access_token") or (data.get("data") or {}).get("access_token")
    assert token, f"no access_token in login response: {data}"
    return token


@pytest.fixture(scope="module")
def admin_token():
    return _login("superadmin@sipro.co.id")


@pytest.fixture(scope="module")
def sales_token():
    return _login("sales@sipro.co.id")


def _h(tok):
    return {"Authorization": f"Bearer {tok}"}


# ---------- Access control / overview ----------
def test_overview_superadmin(admin_token):
    r = requests.get(f"{API}/data-mgmt/overview", headers=_h(admin_token), timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert "entities" in body and "counts" in body and "snapshots" in body
    assert len(body["entities"]) == 15, f"expected 15 entities, got {len(body['entities'])}"
    for ent in body["entities"]:
        assert "key" in ent and "sheet" in ent


def test_overview_forbidden_for_sales(sales_token):
    r = requests.get(f"{API}/data-mgmt/overview", headers=_h(sales_token), timeout=30)
    assert r.status_code == 403


def test_overview_no_auth():
    r = requests.get(f"{API}/data-mgmt/overview", timeout=30)
    assert r.status_code in (401, 403)


# ---------- Template ----------
@pytest.fixture(scope="module")
def template_bytes(admin_token):
    r = requests.get(f"{API}/data-mgmt/template.xlsx", headers=_h(admin_token), timeout=60)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    return r.content


EXPECTED_SHEETS = [
    "PETUNJUK", "DAFTAR NILAI",
    "Pengguna", "Proyek", "Cluster", "Blok", "Tipe Unit", "Unit", "Add-on",
    "Pelanggan", "Vendor", "Subkontraktor", "Mitra", "Material",
    "Tenaga Kerja", "Bagan Akun", "Rekening Bank",
]


def test_template_sheets_and_header(template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), data_only=True)
    for s in EXPECTED_SHEETS:
        assert s in wb.sheetnames, f"missing sheet {s}. Got: {wb.sheetnames}"
    ws = wb["Proyek"]
    # row1 = tech keys, row2 = descriptions, row3 = example
    row1 = [c.value for c in ws[1]]
    row3 = [c.value for c in ws[3]]
    assert "code" in row1 and "name" in row1
    assert row3[0] is not None, "expected example data in row 3"


# ---------- Import dry-run / commit / skip ----------
def test_import_template_dry_run(admin_token, template_bytes):
    files = {"file": ("template.xlsx", template_bytes,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/data-mgmt/import", headers=_h(admin_token),
                      files=files, data={"mode": "upsert", "dry_run": "true"}, timeout=120)
    assert r.status_code == 200, r.text
    body = r.json()
    assert "totals" in body
    assert "entities" in body or "sheets" in body or "results" in body or "reports" in body


def test_import_template_commit_then_skip(admin_token, template_bytes):
    files = {"file": ("template.xlsx", template_bytes,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/data-mgmt/import", headers=_h(admin_token),
                      files=files, data={"mode": "upsert", "dry_run": "false"}, timeout=120)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body.get("totals", {}).get("error", 0) == 0, f"unexpected errors: {body.get('totals')}"

    # Re-import skip mode: all rows should skip
    files = {"file": ("template.xlsx", template_bytes,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = requests.post(f"{API}/data-mgmt/import", headers=_h(admin_token),
                       files=files, data={"mode": "skip", "dry_run": "false"}, timeout=120)
    assert r2.status_code == 200, r2.text
    totals = r2.json().get("totals", {})
    # after commit, second run in skip mode should not insert
    assert totals.get("inserted", 0) == 0, f"expected 0 inserts, got {totals}"


def _build_invalid_workbook():
    """Build workbook with valid header structure but bad data in Proyek sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyek"
    # row 1: technical keys
    ws.append(["code", "name", "location", "status", "members"])
    # row 2: descriptions
    ws.append(["kode", "nama", "lokasi", "status", "anggota"])
    # row 3+: data - invalid enum, missing required
    ws.append(["BAD1", "Nama OK", "Lokasi", "STATUS_TIDAK_VALID", ""])  # bad enum
    ws.append(["", "Tanpa Kode", "Lokasi", "active", ""])  # missing required code
    # Cluster referencing unknown project
    ws2 = wb.create_sheet("Cluster")
    ws2.append(["project_code", "code", "name", "order", "status"])
    ws2.append(["kode proyek", "kode cluster", "nama", "urutan", "status"])
    ws2.append(["PROJECT_TIDAK_ADA", "C1", "Cluster X", 1, "selling"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_invalid_reports_errors(admin_token):
    content = _build_invalid_workbook()
    files = {"file": ("bad.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{API}/data-mgmt/import", headers=_h(admin_token),
                      files=files, data={"mode": "upsert", "dry_run": "true"}, timeout=60)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body.get("totals", {}).get("error", 0) > 0, f"expected errors, got {body}"
    # Check messages are in Indonesian somewhere
    text = json.dumps(body, ensure_ascii=False).lower()
    assert any(k in text for k in ["wajib", "tidak", "harus", "enum", "referensi", "rujukan"]), \
        f"expected Indonesian error keywords, got: {text[:500]}"


# ---------- Export ----------
def test_export_and_reimport_no_errors(admin_token):
    r = requests.get(f"{API}/data-mgmt/export.xlsx", headers=_h(admin_token), timeout=120)
    assert r.status_code == 200, r.text
    content = r.content
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert "Proyek" in wb.sheetnames

    files = {"file": ("export.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = requests.post(f"{API}/data-mgmt/import", headers=_h(admin_token),
                       files=files, data={"mode": "upsert", "dry_run": "true"}, timeout=120)
    assert r2.status_code == 200, r2.text
    totals = r2.json().get("totals", {})
    assert totals.get("error", 0) == 0, f"reimporting export should have 0 errors: {totals}"


# ---------- Backup / snapshot / restore ----------
@pytest.fixture(scope="module")
def backup_bytes(admin_token):
    r = requests.get(f"{API}/data-mgmt/backup.json?include_files=true",
                     headers=_h(admin_token), timeout=120)
    assert r.status_code == 200, r.text
    return r.content


def test_backup_download_meta(backup_bytes):
    payload = json.loads(backup_bytes)
    assert payload["meta"]["format"] == "sipro-backup"
    assert "data" in payload
    assert isinstance(payload["data"], dict) and len(payload["data"]) > 0


def test_snapshot_lifecycle_and_restore(admin_token):
    # create
    r = requests.post(f"{API}/data-mgmt/snapshots", headers=_h(admin_token),
                      data={"label": f"TEST_snap_{int(time.time())}", "include_files": "true"},
                      timeout=120)
    assert r.status_code == 200, r.text
    snap = r.json()
    sid = snap["id"]

    # list
    r = requests.get(f"{API}/data-mgmt/snapshots", headers=_h(admin_token), timeout=30)
    assert r.status_code == 200
    ids = [s["id"] for s in r.json()]
    assert sid in ids

    # download
    r = requests.get(f"{API}/data-mgmt/snapshots/{sid}/download",
                     headers=_h(admin_token), timeout=60)
    assert r.status_code == 200
    payload = json.loads(r.content)
    assert payload["meta"]["format"] == "sipro-backup"

    # restore without confirm -> 400
    r = requests.post(f"{API}/data-mgmt/snapshots/{sid}/restore",
                      headers=_h(admin_token), data={"mode": "merge"}, timeout=30)
    assert r.status_code == 400

    # restore with confirm merge -> success
    r = requests.post(f"{API}/data-mgmt/snapshots/{sid}/restore",
                      headers=_h(admin_token),
                      data={"mode": "merge", "confirm": "RESTORE"}, timeout=180)
    assert r.status_code == 200, r.text
    body = r.json()
    assert "documents" in body
    assert "snapshot_before" in body, f"expected snapshot_before, got: {body}"

    # delete
    r = requests.delete(f"{API}/data-mgmt/snapshots/{sid}",
                        headers=_h(admin_token), timeout=30)
    assert r.status_code == 200


def test_restore_inspect_and_replace_upload(admin_token, backup_bytes):
    # inspect
    files = {"file": ("backup.json", backup_bytes, "application/json")}
    r = requests.post(f"{API}/data-mgmt/restore/inspect", headers=_h(admin_token),
                     files=files, timeout=60)
    assert r.status_code == 200, r.text
    meta = r.json()
    assert meta.get("format") == "sipro-backup"

    # replace with confirm
    files = {"file": ("backup.json", backup_bytes, "application/json")}
    r = requests.post(f"{API}/data-mgmt/restore", headers=_h(admin_token),
                     files=files, data={"mode": "replace", "confirm": "RESTORE"}, timeout=240)
    assert r.status_code == 200, r.text

    # confirm superadmin can still log in
    tok = _login("superadmin@sipro.co.id")
    assert tok
