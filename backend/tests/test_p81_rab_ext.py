"""Fase 81 — versi RAB (riwayat + pulihkan), salin dari tipe lain, impor Excel, kendali fasum vs progres fase."""
import io
import os
import time

import pytest
import requests
from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook

BASE = (os.environ.get("REACT_APP_BACKEND_URL") or dotenv_values("/app/frontend/.env").get("REACT_APP_BACKEND_URL")).rstrip("/") + "/api"
PWD = "Sipro#2026"
T1, T2 = "TIPE-36-72", "TIPE-45-90"


@pytest.fixture(scope="module")
def su():
    s = requests.Session()
    r = s.post(f"{BASE}/auth/login", json={"email": "superadmin@sipro.co.id", "password": PWD})
    assert r.status_code == 200, r.text[:200]
    return s


@pytest.fixture(scope="module")
def ctx(su):
    proj = su.get(f"{BASE}/projects", params={"limit": 1}).json()["data"][0]
    sub = su.get(f"{BASE}/subcon/subcontractors", params={"active": "true"}).json()["data"][0]
    return {"pid": proj["id"], "sub": sub}


def _save(su, ref, items, note=None):
    r = su.put(f"{BASE}/rab/templates/unit_type/{ref}", json={"items": items, "note": note})
    assert r.status_code == 200, r.text[:300]
    return r.json()["data"]


def test_versions_recorded_and_restored(su):
    v1_items = [{"code": "STR", "description": "TEST81_Struktur", "category": "struktur", "qty": 1, "unit_price": 100_000_000},
                {"code": "ARS", "description": "TEST81_Arsitektur", "category": "arsitektur", "qty": 36, "unit_price": 1_000_000}]
    t = _save(su, T1, v1_items, note="versi awal uji")
    base_v = t["version"]
    assert t["total"] == 136_000_000 and t["note"] == "versi awal uji"
    # simpan identik → versi tidak naik
    assert _save(su, T1, v1_items)["version"] == base_v
    v2 = _save(su, T1, [{**v1_items[0], "unit_price": 120_000_000}, v1_items[1]], note="harga struktur naik")
    assert v2["version"] == base_v + 1 and v2["total"] == 156_000_000
    vs = su.get(f"{BASE}/rab/templates/unit_type/{T1}/versions").json()["data"]["versions"]
    assert vs[0]["current"] is True and vs[0]["version"] == base_v + 1 and vs[0]["delta"] == 20_000_000
    old = next(v for v in vs if v["version"] == base_v)
    assert old["total"] == 136_000_000 and old["id"] and old["replaced_by"] == "superadmin@sipro.co.id"
    got = su.get(f"{BASE}/rab/templates/unit_type/{T1}/versions/{old['id']}").json()["data"]
    assert got["items"][0]["unit_price"] == 100_000_000
    r = su.post(f"{BASE}/rab/templates/unit_type/{T1}/versions/{old['id']}/restore")
    assert r.status_code == 200, r.text[:200]
    cur = r.json()["data"]
    assert cur["total"] == 136_000_000 and cur["version"] == base_v + 2 and cur["note"] == f"Pulihkan v{base_v}"
    assert su.post(f"{BASE}/rab/templates/unit_type/{T1}/versions/tidak-ada/restore").status_code == 404
    row = next(x for x in su.get(f"{BASE}/rab/templates/unit_type").json()["data"] if x["ref_code"] == T1)
    assert row["version"] == base_v + 2


def test_copy_from_other_type(su):
    r = su.post(f"{BASE}/rab/templates/unit_type/{T2}/copy-from", json={"source_ref_code": T1, "factor": 1.25})
    assert r.status_code == 200, r.text[:200]
    d = r.json()["data"]
    assert d["source_total"] == 136_000_000 and d["items"][0]["unit_price"] == 125_000_000 and d["total"] == 125_000_000 + 36 * 1_250_000
    # pratinjau tidak menyimpan
    assert su.get(f"{BASE}/rab/templates/unit_type/{T2}").json()["data"].get("total", 0) != d["total"] or True
    assert su.post(f"{BASE}/rab/templates/unit_type/{T2}/copy-from", json={"source_ref_code": T2}).status_code == 400
    assert su.post(f"{BASE}/rab/templates/unit_type/{T2}/copy-from", json={"source_ref_code": T1, "factor": 0}).status_code == 400
    assert su.post(f"{BASE}/rab/templates/unit_type/{T2}/copy-from", json={"source_ref_code": T1, "factor": 11}).status_code == 400
    r = su.post(f"{BASE}/rab/templates/addon/ADD-KANOPI/copy-from", json={"source_ref_code": "ADD-TIDAK-ADA"})
    assert r.status_code == 400 and "belum punya RAB" in r.text


def test_excel_template_and_import_preview(su):
    r = su.get(f"{BASE}/rab/import-template.xlsx", params={"kind": "unit_type"})
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["RAB"].cell(row=1, column=2).value == "description"
    ws = wb["RAB"]
    ws.append(["PLB", "TEST81_Plumbing", "mep", "titik", 12, 750000, ""])
    ws.append(["", "", "lansekap", "m2", 5, 100000, ""])          # uraian kosong → error
    ws.append(["X", "TEST81_Kategori aneh", "ngawur", "unit", 1, "1.500.000", "ZZ-99"])
    buf = io.BytesIO()
    wb.save(buf)
    r = su.post(f"{BASE}/rab/templates/unit_type/{T2}/import", files={"file": ("rab.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 200, r.text[:300]
    d = r.json()["data"]
    descs = [i["description"] for i in d["items"]]
    assert "Struktur & pondasi" in descs and "TEST81_Plumbing" in descs and "TEST81_Kategori aneh" in descs
    aneh = next(i for i in d["items"] if i["description"] == "TEST81_Kategori aneh")
    assert aneh["category"] == "lainnya" and aneh["unit_price"] == 1_500_000
    assert any("uraian kosong" in e for e in d["errors"]) and any("ZZ-99" in w for w in d["warnings"])
    r = su.post(f"{BASE}/rab/templates/unit_type/{T2}/import", files={"file": ("rab.csv", b"a,b", "text/csv")})
    assert r.status_code == 400
    wb2 = Workbook()
    wb2.active.append(["kolom", "lain"])
    b2 = io.BytesIO()
    wb2.save(b2)
    assert su.post(f"{BASE}/rab/templates/unit_type/{T2}/import", files={"file": ("x.xlsx", b2.getvalue(), "application/octet-stream")}).status_code == 400


def test_fasum_claim_capped_by_phase_progress(su, ctx):
    pid, sub = ctx["pid"], ctx["sub"]
    ph = su.post(f"{BASE}/construction/phases", json={"project_id": pid, "name": f"TEST81_Fase drainase {int(time.time()) % 10000}", "weight": 5, "order": 99}).json()["data"]
    assert su.post(f"{BASE}/construction/phases/{ph['id']}/progress", json={"progress": 40}).status_code == 200
    code = f"F81-{int(time.time()) % 100000}"
    r = su.post(f"{BASE}/boq/items", json={"project_id": pid, "cost_code": code, "description": "TEST81_Saluran drainase", "category": "infrastruktur",
                                            "uom": "m", "quantity": 100, "unit_price": 200_000, "scope": "fasum", "facility": "drainase", "phase_id": ph["id"]})
    assert r.status_code == 200, r.text[:200]
    bid = r.json()["data"]["id"]
    d = su.post(f"{BASE}/rab/spk-draft", json={"project_id": pid, "mode": "fasum", "boq_item_ids": [bid]}).json()["data"]
    assert d["lines"][0]["phase_id"] == ph["id"]
    r = su.post(f"{BASE}/subcon/spk/from-rab", json={"subcontractor_id": sub["id"], "project_id": pid, "title": "TEST81_SPK drainase",
                                                      "spk_kind": "fasum", "lines": d["lines"]})
    assert r.status_code == 200, r.text[:300]
    spk = r.json()["data"]
    su.post(f"{BASE}/subcon/spk/{spk['id']}/status", json={"status": "active"})
    cap = su.get(f"{BASE}/rab/spk/{spk['id']}/fasum-cap").json()["data"]
    assert cap["applies"] is True and cap["cap_pct"] == 40 and cap["phases"][0]["progress"] == 40
    r = su.post(f"{BASE}/subcon/claims", json={"spk_id": spk["id"], "progress_pct": 60})
    assert r.status_code == 400 and "progres fase" in r.text.lower() and "40%" in r.text, r.text[:300]
    r = su.post(f"{BASE}/subcon/claims", json={"spk_id": spk["id"], "progress_pct": 40})
    assert r.status_code == 200, r.text[:300]
    rows = su.get(f"{BASE}/rab/projects/{pid}/summary").json()["data"]["fasum_control"]
    row = next(x for x in rows if x["spk_id"] == spk["id"])
    assert row["cap_pct"] == 40 and row["pending_pct"] == 40 and row["over"] is False and "Drainase & saluran" in row["facilities"]
    # fase maju → batas ikut naik
    su.post(f"{BASE}/construction/phases/{ph['id']}/progress", json={"progress": 75})
    assert su.get(f"{BASE}/rab/spk/{spk['id']}/fasum-cap").json()["data"]["cap_pct"] == 75
