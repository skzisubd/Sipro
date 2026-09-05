"""Fase 81 — perluasan RAB tipe/add-on: riwayat VERSI (pulihkan), SALIN dari tipe lain (dengan faktor),
IMPOR Excel (template + pratinjau tervalidasi). Semua penyimpanan tetap lewat `rab_engine.save_template`
sehingga versi tercatat satu jalur."""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import rab_engine as re_
from db import db
from reference import values as ref_values

IMPORT_COLUMNS = [("code", "Kode"), ("description", "Uraian pekerjaan (wajib)"), ("category", "Kategori"),
                  ("uom", "Satuan"), ("qty", "Volume"), ("unit_price", "Harga satuan (Rp)"), ("step_code", "Kode langkah jadwal")]
MAX_FACTOR = 10.0


# ============================================================ versi
async def list_versions(org: str, kind: str, ref_code: str) -> dict:
    cur = await re_.get_template(org, kind, ref_code)
    hist = await db.rab_template_versions.find({"org_id": org, "kind": kind, "ref_code": ref_code},
                                               {"_id": 0}).sort("version", -1).to_list(200)
    rows = [{"id": None, "version": int(cur.get("version") or (1 if cur.get("items") else 0)), "current": True,
             "total": re_._i(cur.get("total")), "items_count": len(cur.get("items") or []),
             "saved_by": cur.get("updated_by"), "saved_at": cur.get("updated_at"), "note": cur.get("note")}]
    rows += [{"id": h["id"], "version": h["version"], "current": False, "total": re_._i(h.get("total")),
              "items_count": len(h.get("items") or []), "saved_by": h.get("saved_by"), "saved_at": h.get("saved_at"),
              "replaced_by": h.get("replaced_by"), "replaced_at": h.get("replaced_at"), "note": h.get("note")} for h in hist]
    for i, r in enumerate(rows):
        r["delta"] = r["total"] - rows[i + 1]["total"] if i + 1 < len(rows) else None
    return {"kind": kind, "ref_code": ref_code, "versions": rows if cur.get("items") or hist else []}


async def get_version(org: str, kind: str, ref_code: str, vid: str) -> dict:
    v = await db.rab_template_versions.find_one({"org_id": org, "kind": kind, "ref_code": ref_code, "id": vid}, {"_id": 0})
    if not v:
        raise LookupError("Versi RAB tidak ditemukan.")
    return v


async def restore_version(org: str, kind: str, ref_code: str, vid: str, actor: str) -> dict:
    v = await get_version(org, kind, ref_code, vid)
    return await re_.save_template(org, kind, ref_code, v["items"], actor, note=f"Pulihkan v{v['version']}")


# ============================================================ salin dari tipe/add-on lain
async def copy_items(org: str, kind: str, ref_code: str, source_ref: str, factor: float = 1.0) -> dict:
    if kind not in ("unit_type", "addon"):
        raise ValueError("Jenis template harus unit_type atau addon.")
    if source_ref == ref_code:
        raise ValueError("Sumber salinan harus tipe/add-on yang berbeda.")
    try:
        factor = float(factor if factor is not None else 1.0)
    except (TypeError, ValueError):
        raise ValueError("Faktor harga harus angka.")
    if not 0 < factor <= MAX_FACTOR:
        raise ValueError(f"Faktor harga harus di antara 0 dan {MAX_FACTOR:g} (1 = harga sama).")
    src = await re_.get_template(org, kind, source_ref)
    if not src.get("items"):
        raise ValueError(f"Sumber '{source_ref}' belum punya RAB — tidak ada yang bisa disalin.")
    items = re_.normalize_items([{**it, "unit_price": re_._i(re_._i(it.get("unit_price")) * factor)} for it in src["items"]])
    return {"kind": kind, "ref_code": ref_code, "source_ref": source_ref, "factor": factor, "source_total": re_._i(src.get("total")),
            "items": items, "total": sum(i["amount"] for i in items)}


# ============================================================ impor Excel
def import_workbook(kind: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RAB"
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    for c, (key, label) in enumerate(IMPORT_COLUMNS, start=1):
        ws.cell(row=1, column=c, value=key).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=c).fill = head_fill
        ws.cell(row=2, column=c, value=label).font = Font(italic=True, color="666666")
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 22 if key != "description" else 40
    example = ["STR", "Struktur & pondasi", "struktur", "unit", 1, 150000000, ""] if kind == "unit_type" \
        else ["R01", "Rangka + atap kanopi", "finishing", "unit", 1, 8000000, ""]
    for c, v in enumerate(example, start=1):
        ws.cell(row=3, column=c, value=v)
    cats = ",".join(ref_values("work_category"))
    dv = DataValidation(type="list", formula1=f'"{cats}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("C3:C500")
    ws.freeze_panes = "A3"
    info = wb.create_sheet("PETUNJUK")
    for i, t in enumerate([
        "Baris 1 = kunci kolom (jangan diubah), baris 2 = keterangan, baris 3 dst = data RAB.",
        "Uraian wajib; Volume > 0; Harga satuan angka bulat rupiah tanpa titik.",
        f"Kategori pilih dari: {', '.join(ref_values('work_category'))}.",
        "Kode langkah jadwal (opsional, hanya RAB tipe unit) harus ada di template jadwal pembangunan.",
        "Hasil unggah = PRATINJAU di editor RAB; tekan Simpan RAB untuk menyimpan (versi lama tersimpan di riwayat).",
    ], start=1):
        info.cell(row=i, column=1, value=t)
    info.column_dimensions["A"].width = 110
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _known_steps(org: str) -> set:
    out = set()
    async for t in db.build_templates.find({"org_id": org}, {"_id": 0, "steps.code": 1}):
        out.update(s.get("code") for s in t.get("steps") or [] if s.get("code"))
    return out


async def parse_import(org: str, kind: str, content: bytes) -> dict:
    if kind not in ("unit_type", "addon"):
        raise ValueError("Jenis template harus unit_type atau addon.")
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        raise ValueError("Berkas bukan Excel .xlsx yang sah.")
    ws = wb["RAB"] if "RAB" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet RAB kosong.")
    header = [str(h or "").strip().lower() for h in rows[0]]
    keys = [k for k, _ in IMPORT_COLUMNS]
    if "description" not in header:
        raise ValueError("Baris 1 harus memuat kunci kolom template (minimal 'description').")
    idx = {k: header.index(k) for k in keys if k in header}
    cats = set(ref_values("work_category"))
    steps = await _known_steps(org) if kind == "unit_type" else set()
    items, errors, warnings = [], [], []
    for rn, row in enumerate(rows[1:], start=2):
        if rn == 2 and idx.get("description") is not None and str(row[idx["description"]] or "").lower().startswith("uraian"):
            continue
        get = lambda k: (row[idx[k]] if k in idx and idx[k] < len(row) else None)  # noqa: E731
        if all(v in (None, "") for v in row):
            continue
        desc = str(get("description") or "").strip()
        if not desc:
            errors.append(f"Baris {rn}: uraian kosong.")
            continue
        try:
            qty = float(get("qty") if get("qty") not in (None, "") else 1)
            price = float(str(get("unit_price") or 0).replace(".", "").replace(",", "").replace("Rp", "").strip() or 0)
        except (TypeError, ValueError):
            errors.append(f"Baris {rn}: volume/harga satuan bukan angka.")
            continue
        if qty <= 0 or price < 0:
            errors.append(f"Baris {rn}: volume harus > 0 dan harga satuan ≥ 0.")
            continue
        cat = str(get("category") or "lainnya").strip().lower()
        if cat not in cats:
            warnings.append(f"Baris {rn}: kategori '{cat}' tidak dikenal → 'lainnya'.")
            cat = "lainnya"
        step = str(get("step_code") or "").strip() or None
        if step and kind != "unit_type":
            warnings.append(f"Baris {rn}: kode langkah diabaikan untuk RAB add-on.")
            step = None
        if step and steps and step not in steps:
            warnings.append(f"Baris {rn}: langkah '{step}' tidak ada di template jadwal — tidak akan masuk lingkup.")
        items.append({"code": str(get("code") or "").strip() or f"R{len(items) + 1:02d}", "description": desc, "category": cat,
                      "uom": str(get("uom") or "unit").strip(), "qty": qty, "unit_price": int(round(price)), "step_code": step})
    norm = re_.normalize_items(items)
    return {"items": norm, "total": sum(i["amount"] for i in norm), "rows": len(norm), "errors": errors, "warnings": warnings}
